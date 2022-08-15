import openpyxl as op
import openpyxl as ks
import pandas as pd

def show_attr(choice, db, column):
    k = 1
    for i in db:
        if db[i][column].lower() == choice.lower():
            print(k,end=' ')
            k+=1
            print(db[i])

def filter_cars(choice,db,column):
    cars = {}
    k = 0
    for i in db:
        if db[i][column].lower() == choice.lower():
            k += 1
            cars[k] = [i,*db[i]]
    return cars

def set_status(db,status,**attrs):
    brand, model, kpp = attrs.get("brand"), attrs.get("model"), attrs.get("kpp")
    car = None
    filter_brand = filter_model = filter_kpp = {}
    if brand:
        filter_brand = {k: v for k, v in db.items() if v[0] == brand}
    if model:
        filter_model = {k: v for k, v in db.items() if v[1] == model}
    if kpp:
        filter_kpp = {k: v for k, v in db.items() if v[2] == kpp}
    if all((brand, model, kpp)):
        for k in filter_brand:
            if k in filter_model and k in filter_kpp:
                car = filter_brand[k]
                break
    elif brand and model:
        for k in filter_brand:
            if k in filter_model:
                car = filter_brand[k]
                break
    elif model and kpp:
        for k in filter_model:
            if k in filter_kpp:
                car = filter_model[k]
                break
    elif brand and kpp:
        for k in filter_model:
            if k in filter_brand:
                car = filter_model[k]
                break
    if not car: return "Такой машины нет"
    db[k][-1] = status

def delete_line(file,dline):
    with open(file) as f:
        lines=f.readlines()
    with open(file,'w') as f:
        for line in lines:

            if ' '.join(line.replace('\n','').split()[:-1])==' '.join(dline.split()[:-2]):
                continue
            f.write(line)

def write(file,cars):
    with open(file,'w') as f:
        for car in cars:
            f.write(' '.join(cars[car][1:])+'\n')

def show_choice(choice, allcars):
    if choice == 0:
        choice_dict = {0: "mercedes-benz", 1: "bmw", 2: "toyota", 3: "honda", 4: "volkswagen"}
        text = "brands : mercedes-benz = 1, bmw = 2, toyota = 3, honda = 4, volkswagen = 5 :\n"
        value = int(input(text)) - 1
        while value not in range(5):
            print("invalid choice")
            value = int(input(text)) - 1
        show_attr(choice_dict[value], allcars, 0)
        car_id = int(input())
        cars_by_model = {}
        k = 1
        for i in allcars:
            if allcars[i][0].lower() == choice_dict[value].lower():
                cars_by_model[k] = allcars[i]
                k+=1
        car_choice_list = cars_by_model[car_id]
        wb = op.load_workbook('auto_info.xlsx')
        wb.active = 0
        sheet = wb.active
        for cell in sheet['B']:
            if cell.value == car_choice_list[1]:
                print(cell.value)
        allcars[car_id] = 'sold'
        print(car_choice_list)
        car_choice_list[-1] = 'sold'
        sold = open("solt-cars.txt", 'w')
        sold.write(str(car_choice_list))
        sold.close()
    elif choice == 1:
        print(set([allcars[i][1] for i in allcars]))
        text = "model:\n"
        value = input(text)
        show_attr(value, allcars, 1)
    elif choice == 2:
        choice_dict = {1: "automate", 2: "mechanical"}
        text = "kpp: automate = 1, mechanical = 2\n"
        value = int(input(text))
        while value not in range(1, 3):
            print("invalid choice")
            value = int(input(text))
        show_attr(choice_dict[value], allcars, 2)
    elif choice == 3:
        choice_dict = {1: "microbus", 2: "sedan", 3: "jeep", 4: "universal", 5: "kupe"}
        text = "body: microbus = 1, sedan = 2, jeep = 3, universal = 4, kupe = 5\n"
        value = int(input(text))
        while value not in range(1, 6):
            print("invalid choice")
            value = int(input(text))
        show_attr(choice_dict[value], allcars, 3)
    elif choice == 4:
        print(set([allcars[i][4] for i in allcars]))
        text = "price:\n"
        value = input(text)
        show_attr(value, allcars, 4)
    elif choice == 5:
        choice_dict = {1: "in stock", 2: "booked", 3: "in service", 4: "sold"}
        text = "status: in stock = 1, booked = 2, in service = 3, sold = 4\n"
        value = int(input(text))
        while value not in range(1, 6):
            print("invalid choice")
            value = int(input(text))
        show_attr(choice_dict[value], allcars, 5)

def get_choice():
    choice = int(input("brand = 1, model = 2, kpp = 3, body = 4, price = 5, status = 6:\n")) - 1
    while choice not in range(0, 6):
        print("invalid choice")
        choice = int(input("brand = 1, model = 2, kpp = 3, body = 4, price = 5, status = 6:\n")) - 1
    return choice

def get_db(file):
    wb = op.load_workbook(file)
    sheets = wb.sheetnames
    for sheet in sheets:
        print(sheet)
    sheet = wb.active
    allcars = {}
    id = 1
    for row in sheet.rows:
        arraytmp = []
        for cell in row:
            arraytmp.append(str(cell.value))
        allcars[id] = arraytmp
        id += 1
    return allcars

def set_db(file,allcars):
    wb = op.load_workbook(file)
    sheet=wb.active
    row=1
    column=1
    for car in allcars:
        for attr in allcars[car]:
            sheet.cell(row=row,column=column).value=attr
            column+=1
        column=1
        row+=1
    wb.save(file)

def counter_car(x):
    allcars=get_db('auto_info.xlsx')
    count=0
    for i in allcars:
        if allcars[i][-1]==x:
            count+=1
    return count

def set_salary(salary):
    with open ('repairman_salary.txt')as f:
        old_salary=f.read()
    with open('repairman_salary.txt','w')as f:
        if old_salary:
            f.write(str(int(old_salary)+salary))
        else:
            f.write(str(salary))

def menu_director():
    while True:
        print(' 1.весь список автомобилей \n 2.Количество проданных автомобилей \n 3.Показать автомобиль с самым максимальным количествои продаж'
              ' \n 4.Показать автомобиль с самым минимальным количествои продаж \n 5.Показать автомобиль,который требует больше всего сервисного обслуживания  \n 6.Показать самый дорогой автомобиль  \n 7.Показать самый дешевый автомобиль \n 8.выход')
        c=int(input('Пожалуйста наберите номер меню для работы с программой, если закончили, то наберите 8:'))
        if c==1:
            db_xl = get_db('auto_info.xlsx')
            for i in db_xl:
                print(i, ':', db_xl[i])
        elif c==2:
            show_attr("in sold", get_db('sold-cars.xlsx'), 3)
        elif c==3:
            wd = pd.read_excel('sold-cars.xlsx')
            brand = []
            model = []
            price = []
            a = []
            for i in wd['price']:
                a.append(i)
            b = max(a)
            for i in range(len(wd['price'])):
                if wd['price'][i] == b:
                    brand.append(wd['brand'][i])
                    model.append(wd['model'][i])
                    price.append(wd['price'][i])
            print(pd.DataFrame({
                'Brand': brand,
                'Model': model,
                'Count': price,
            }))
        elif c==4:
            wd = pd.read_excel('sold-cars.xlsx')
            brand = []
            model = []
            price = []
            a = []
            for i in wd['price']:
                a.append(i)
            b = min(a)
            for i in range(len(wd['price'])):
                if wd['price'][i] == b:
                    brand.append(wd['brand'][i])
                    model.append(wd['model'][i])
                    price.append(wd['price'][i])
            print(pd.DataFrame({
                'Brand': brand,
                'Model': model,
                'Count': price,
            }))
        elif c==5:
            book = ks.open('auto_info.xlsx', read_only=True)
            sheet = book.active
            for row in range(2, 3):
                print(sheet[row][0].value, sheet[row][1].value)
        elif c == 6:
            wd = pd.read_excel('auto_info.xlsx')
            max1 = wd['price'][0]
            index1 = 0
            for i in range(len(wd['price'])):
                if max1 < wd['price'][i]:
                    max1 = int(wd['price'][i])
                    index1 = i
            d = {
                'Brand': wd['brand'][index1],
                'Model': wd['model'][index1],
                'Price': wd['price'][index1],
            }
            print()
            print(pd.DataFrame([d]))
            print()
        elif c == 7:
            wd = pd.read_excel('auto_info.xlsx')
            min1 = int(wd['price'][0])
            index1 = 0
            for i in range(len(wd['price'])):
                if min1 >wd['price'][i]:
                    min1 = int(wd['price'][i])
                    index1 = i
            d = {
                'Brand': wd['brand'][index1],
                'Model': wd['model'][index1],
                'Price': wd['price'][index1],
            }
            print()
            print(pd.DataFrame([d]))
            print()
        elif c==8:
            break

def menu_seller():
    print('Приветствую, дорогой Продавец!')
    allcars = get_db('auto_info1.xlsx')
    while True:
        print(' 1.весь список автомобилей \n 2.поиск автомобиля \n 3.отчет по автомобилям \n 4.заказ автомобиля \n 5.список проданных автомобилей \n 6.вернуть купленное авто \n 7.выход')
        c=int(input('Пожалуйста наберите номер меню для работы с программой, если закончили, то наберите 7:'))
        if c==1:
            for i in allcars:
                print(i, ':', allcars[i])
        elif c==2:
            show_choice(get_choice(), allcars)
        elif c==3:
            print('in stock:',counter_car('in stock'))
            print('sold:',counter_car('sold'))
            print('booked:',counter_car('booked'))
            print('in service:',counter_car('in service'))
        elif c==4:
            print('Пожалуйста напишите какое авто вы бы хотели заказать:')
            show_attr("in stock", allcars, 5)
            n = input('введите цифру машины: ')
            while not n.isdigit():
                n = input('введите цифру: ')
                break
            cars = filter_cars('in stock', allcars, 5)
            g = cars[int(n)]
            g[-1]='sold'
            allcars[cars[int(n)][0]][-1]='sold'
            set_db('auto_info1.xlsx',allcars)
            write('solt-cars.txt',filter_cars('sold',allcars,5))
            price=int(g[5])
            print(f'''
            Цена за {g[1]} {g[2]} : {price}$.
            Сумма налога составила: 1% или {price*0.01}$.
            Сумма комиссий продавцу за продажу составила: 0.5% или {price*0.005}$. 
            Итого окончательная цена: {price+price*0.01+price*0.005}$.''')
        elif c == 5:
            show_attr("sold", allcars, 5)
        elif c == 6:
            print('Пожалуйста напишите какое авто вы бы хотели вернуть:')
            show_attr('sold',allcars,5)
            n = input('введите цифру машины: ')
            while not n.isdigit():
                n = input('введите цифру: ')
                break
            cars=filter_cars('sold',allcars,5)
            g = cars[int(n)]
            g[-1]='in stock'
            allcars[g[0]][-1] = 'in stock'
            set_db('auto_info1.xlsx', allcars)
            write('returned.txt',filter_cars('in stock',allcars,5))
            delete_line('solt-cars.txt',' '.join(g[1:]))
        elif c==7:
            break

def menu_repairman():
	print('Приветствую,дорогой Механик!')
	allcars = get_db('auto_info1.xlsx')
	print(allcars)
	while True:
		print(' 1.добавить автомобиль на обслуживание \n 2.показать список всех автомобилей на обслуживании \n 3.обслужить автомобиль \n 4.показать обслуженные автомобили \n 5.показать мой заработок \n 6.выход ')
		c = int(input('Пожалуйста наберите номер меню для работы с программой, если закончили, то наберите 6:'))
		if c==1:
			b=input('введите бренд автомобиля:')
			m=input('введите модель автомобиля:')
			r=input('введите причину обслуживания:')
			sc = open('in_service.txt', 'w',encoding='utf-8' )
			sc.write(" ".join([b, m, 'по причине :',r]))
			sc.close()
			res=set_status(allcars,'in service',brand=b,model=m)
			print(res)
			set_db('auto_info1.xlsx',allcars)
			set_salary(5000)
		elif c==2:
			show_attr("in service", allcars, 5)
		elif c==3:
			print('Пожалуйста напишите какое авто вы бы хотели обслужить:')
			show_attr('in service', allcars, 5)
			n = input('введите цифру машины: ')
			while not n.isdigit():
				n = input('введите цифру: ')
				break
			cars = filter_cars('in service', allcars, 5)
			g = cars[int(n)]
			g[-1] = 'served'
			allcars[cars[int(n)][0]][-1] = 'served'
			set_db('auto_info1.xlsx', allcars)
			write('served.txt', filter_cars('served', allcars, 5))
		elif c==4:
			show_attr("served", allcars, 5)
		elif c==5:
			with open('repairman_salary.txt') as f:
				print(f.read(),'$',sep='')
		elif c==6:
			break


