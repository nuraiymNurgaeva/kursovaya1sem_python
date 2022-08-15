import openpyxl as xl




print('Добро пожаловать в автосалон Nitro')
while True:
    wb = xl.load_workbook('accounts.xlsx')
    sheet = wb.active
    a = []
    b = []
    c = []
    for cell in list(sheet.columns)[0][1:]:
        a.append(cell.value)
    for cell in list(sheet.columns)[1][1:]:
        b.append(cell.value)
    for cell in list(sheet.columns)[2][1:]:
        c.append(str(cell.value))

    ak = input('Введите аккаунт: ')
    if ak not in a:
        print('Извините, но мы не нашли такой тип аккаунта, пожалуйста повторите.')

    elif ak in a:
        print('Проверка аккаунта прошла успешно')

        while True:
            login = input('Введите логин')
            password = input('Введите пороль')
            if login not in b:
                print('Неверный логин, введите еще раз')
            elif password not in c:
                print('Неверный пороль,введите еще раз')
            elif password not in c and login not in b:
                print('неверный логин и пороль,введите еще раз')
            elif login in b[1] and password in c[1]:
                    print('Приветствую дорогой,Продавец')

                    def menu():
                        print('1.список автомобилей')
                        print('2.поиск автомобилей')
                        print('3.отчет по автомобилям')
                        print('4.заказ автомобилей')
                        print('5.список купленных автомобилей')
                        print('6.вернуть проданный автомобиль')
                        print('7.какой автомобиль вы хотели б вернуть?')
                        print('8.выход')

                        while True:
                            click=input('Выберите пункт')
                            if click=='1':
                                book = xl.open('autos.xlsx', read_only=True)
                                sheet = book.active

                                for row in range(1, sheet.max_row + 1):
                                    brand = sheet[row][0].value
                                    model = sheet[row][1].value

                                    print(brand, model)

                                while True:
                                    c1=input('Введите Back,чтобы перейти к меню')
                                    if c1==('back'):
                                        menu()
                                    else:
                                        print('Не верно,Введите back')
                            elif click=='2':
                                print('1).Модель')
                                print('2).Марка')
                                print('3).Кузов')
                                print("4).АКПП")
                                book = xl.open('autos.xlsx', read_only=True)
                                sheet = book.active
                                while True:
                                    click=input('Выберите пункт:')
                                    if click=='1':
                                        m1=input('введите модель:')
                                        a=[]
                                        for t in range(1, sheet.max_row + 1):
                                            model = sheet[t][1].value
                                            brand = sheet[t][0].value

                                            if model==m1:
                                                a.append([brand,model])
                                        for i1 in a:
                                            print(*i1)





                                        while True:    
                                            c1=input('Введите Back,чтобы перейти к меню')
                                            if c1==('back'):
                                                menu()
                                            else:
                                                print('Не верно,Введите back')
                                    elif click=='2':
                                        m1 = input('введите Бренд:')
                                        a=[]
                                        for t in range(1, sheet.max_row + 1):
                                            model = sheet[t][1].value
                                            brand = sheet[t][0].value

                                            if brand == m1:
                                                a.append([brand, model])
                                        for i1 in a:
                                            print(*i1)


                                        while True:    
                                            c1=input('Введите Back,чтобы перейти к меню')
                                            if c1==('back'):
                                                menu()
                                    elif click=='3':
                                        m1 = input('введите Бренд')
                                        a = []
                                        for t in range(1, sheet.max_row + 1):
                                            model = sheet[t][1].value
                                            brand = sheet[t][0].value
                                            body = sheet[t][2].value


                                            if brand == m1:
                                                a.append([brand, model,body])
                                        for i1 in a:
                                            print(*i1)


                                    elif click=='4':
                                        m1 = input('введите Бренд')
                                        a = []
                                        for t in range(1, sheet.max_row + 1):
                                            model = sheet[t][1].value
                                            brand = sheet[t][0].value
                                            kpp = sheet[t][3].value

                                            if brand == m1:
                                                a.append([brand, model, kpp])


                                        while True:    
                                            c1=input('Введите Back,чтобы перейти к меню')
                                            if c1==('back'):
                                                menu()
                                    
                                    else:
                                         print('не верно')
                            elif click=='3':
                                book = xl.open('autos.xlsx', read_only=True)

                                sheet = book.active
                                print('Имеющиеся машины:')
                                for row in range(1, sheet.max_row + 1):
                                    brand = sheet[row][0].value
                                    model = sheet[row][1].value
                                    price = sheet[row][4].value

                                    print(brand, model, price)
                                print('список проданных:')
                                book = xl.open('solt-cars.txt.xlsx', read_only=True)
                                sheet = book.active
                                for row in range(1, sheet.max_row + 1):
                                    print(sheet[row][0].value, sheet[row][1].value)
                                print('на ремонте:')
                                book = xl.open('sh.xlsx', read_only=True)
                                sheet = book.active
                                for row in range(1, sheet.max_row + 1):
                                    print(sheet[row][0].value, sheet[row][1].value)
                            elif click=='4':
                                while True:
                                    q=input('пожалуйста напишите ккакой автомобиль вы бы хотели').split()
                                workbook = xlsxwriter.Workbook('solt-cars.txt.xlsx')
                                wd=workbook.add_worksheet()

                                wd.write(1,0,q[0])
                                wd.write(1,1,q[1])
                                workbook.close()

                                if q=='back':
                                    menu()
                            elif click=='8':
                                print('Завершение Работы')
                                break


                    menu()    
            elif login in b[0] and c[0]:
                print('Приветствую дорогой,Директор')

                def menu1():
                    print('Наберите номер для работы с пргораммой:')
                    print('Пункт-1.Список автомобилей доступных к продаже')
                    print('Пункт-2.Показать количество всех проданных авто')
                    print('Пункт-3.Показать авто с самым максимальным кол продаж')
                    print('Пункт-4.Показать авто с самым минимальным кол продаж')
                    print('Пункт-5.Показать самый дорогой авто')
                    print('Пункт-6.показать самый дешевый авто')
                    print('Пункт-7.Показать авто требуюший сервисного обслуживание')

                menu1()
                while True:
                    click1 = input('Выберете пункт:')
                    if click1 == '1':
                        a1 = ['mers124', 'audi 100', 'bmw x5']
                        print(a1)

                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == 'back':
                            menu1


                    elif click1 == '2':
                        prod_a = ['mers', 'kia']
                        print(prod_a)
                        s=input('Введите Back,чтобы перейти к меню')
                        if s=='back':
                            menu1()
                    elif click1 == '3':
                        max_prod = ['audi']
                        print(max_prod)
                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == ('back'):
                            menu1()
                    elif click1 == '4':
                        min_prod = ['kia']
                        print(min_prod)
                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == ('back'):
                            menu1()
                    elif click1 == '5':
                        dor_avto = ['mers']
                        print(dor_avto)
                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == ('back'):
                            menu1()
                    elif click1 == '6':
                        desh = ['kia']
                        print(desh)
                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == ('back'):
                            menu1()

                    elif click1 == '7':
                        serv = ['honda']
                        print(serv)
                        c1 = input('Введите Back,чтобы перейти к меню')
                        if c1 == ('back'):
                            menu1()
                    


                    elif click1 == '8':
                        print('Программа завершена, мы будем рады вашему возврашению')
                menu1()
            elif login in b[2] and password in c[2]:
                print('Приветствую, дорогой Механик!')


                def menu2():
                    print('Наберите номер меню для работы с программой:')
                    print('1.Забрать автомобиль на обслуживание')
                    print('2.Показать список всех автомобилей на обслуживание')
                    print('3.Обслужить автомобиль')
                    print('4.Показать обслуженные автомобили')
                    print('5.Показать мой заработок')
                menu2()
                while True:
                    click1 = input('Выберете номер:')
                    if click1 == '1':
                        user_brand = input('Выберете брэнд автомобиля:')
                        user_model = input('Введите модель автомобиля:')
                        reason = input('Введите причину обслуживания:')
                        cars = []
                        user_car = [user_brand, user_model]
                        cars.append('на обслуживании=',user_car)
                        print(cars)
                        in_service=[]
                        in_service.append(user_car,'по причине',reason)
                        print(in_service)
                    elif click1 =='2':
                        for i in cars:
                            if 'на обслуживании=' in i:
                                print(i)
                        ordered_equipment=['выпрессовщик ','станок для проточки тормозных дисков','пескоструйный станок','станкок для обдирки и заточки инструмента']
                        print('список оборудований для доставки:',ordered_equipment,)
                    elif click1=='3':
                        nazv=input('Введите название автомобиля, которое было обслужено(брэнд,модель авто):')
                        if nazv in in_service:
                            in_service.remove(nazv)
                        else:
                            print('Такой автомобиль не найден , попробуйте ввести название автомобиля еще раз:')
                        served_cars=[]
                        served_cars.append(nazv)
                        cars.clear()
                    elif click1=='4':
                        print('список,обслуженных автомобилей:',served_cars)
                    elif click1=='5':
                        n=served_cars.count()
                        money=n*1000
                        print(money,'сом')
                    elif click1=='6':
                        print('Программа завершена, мы будем рады вашему возвращению!')
                menu2()        
